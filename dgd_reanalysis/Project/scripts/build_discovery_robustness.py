#!/usr/bin/env python3
"""Dataset-level robustness and expanded-comparator analyses.

This script does not rebuild raw single-cell objects. It re-expresses the
existing primary-core outputs at dataset-level inference units, stratifies
candidate support by species, tests leave-one-dataset stability, evaluates
matched null gene sets, and uses Allen whole-mouse-brain taxonomy markers as
an independent four-system comparator layer.
"""

from __future__ import annotations

import math
import os
import re
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats


ROOT = Path(__file__).resolve().parents[2]
RESULTS = ROOT / "Project/results"
ALLEN_XLSX = (
    ROOT
    / "External_Data/Allen_Institute/WMB_taxonomy_cluster_annotation_CCN202307220.xlsx"
)

DATASET_DELTAS = RESULTS / "primary_core_mgi_ortholog_formal_rank_dataset_deltas.tsv.gz"
GENE_SUMMARY = RESULTS / "primary_core_mgi_ortholog_formal_rank_gene_summary.tsv"
CANDIDATE_TIERS = RESULTS / "primary_core_manuscript_candidate_tiers.tsv"
CONFIGURATION = RESULTS / "primary_core_transcriptomic_configuration_primary_contrasts.tsv"
MODULE_SUMMARY = RESULTS / "primary_core_niche_circuit_module_formal_summary.tsv"
MODULE_GENE_SCORES = RESULTS / "primary_core_niche_circuit_module_formal_gene_scores.tsv"
MODULE_GENE_SETS = RESULTS / "primary_core_niche_circuit_module_gene_sets.tsv"

OUT_CONFIG_DATASETS = RESULTS / "dgd_dataset_level_configuration.tsv"
OUT_CONFIG_SUMMARY = RESULTS / "dgd_dataset_level_configuration_summary.tsv"
OUT_MODULE_INFERENCE = RESULTS / "dgd_module_level_inference.tsv"
OUT_SPECIES = RESULTS / "dgd_species_stratified_candidates.tsv"
OUT_SPECIES_SHARED = RESULTS / "dgd_species_shared_support.tsv"
OUT_LODO = RESULTS / "dgd_candidate_leave_one_dataset_out.tsv"
OUT_LODO_SUMMARY = RESULTS / "dgd_candidate_lodo_summary.tsv"
OUT_NULL = RESULTS / "dgd_matched_null_summary.tsv"
OUT_MODULE_SENSITIVITY = RESULTS / "dgd_module_leave_one_gene_out.tsv"
OUT_MODULE_OVERLAP = RESULTS / "dgd_module_gene_overlap.tsv"
OUT_ALLEN_CLUSTERS = RESULTS / "dgd_allen_comparator_clusters.tsv"
OUT_ALLEN_MODULES = RESULTS / "dgd_allen_comparator_module_markers.tsv"
OUT_ALLEN_CANDIDATES = RESULTS / "dgd_allen_candidate_marker_prevalence.tsv"
OUT_FIGURE = RESULTS / "dgd_robustness_and_comparators.png"
OUT_REPORT = RESULTS / "dgd_robustness_and_comparators.md"

STRICT_DATASETS = {
    "GSE104323": ("mouse", "dentate"),
    "GSE95752": ("mouse", "dentate"),
    "GSE292261": ("mouse", "dentate"),
    "GSE214309": ("mouse", "dentate"),
    "GSE122357": ("mouse", "cerebellum"),
    "GSE312658": ("mouse", "cerebellum"),
    "GSE165657": ("human", "cerebellum"),
    "GSE186538": ("human", "dentate"),
    "GSE325391": ("human", "dentate"),
    "GSE268609": ("human", "dentate"),
}

TIER1_LABEL = "Tier 1 core convergent program"
TIER2_LABEL = "Tier 2 high-confidence wiring/synaptic executor"
RNG = np.random.default_rng(205920)
N_BOOTSTRAP = 10_000
N_NULL = 10_000

os.environ.setdefault("MPLCONFIGDIR", str(RESULTS / "mplconfig"))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns


def bh_adjust(values: pd.Series) -> pd.Series:
    p = pd.to_numeric(values, errors="coerce").to_numpy(dtype=float)
    out = np.full(len(p), np.nan)
    valid = np.isfinite(p)
    if not valid.any():
        return pd.Series(out, index=values.index)
    pv = p[valid]
    order = np.argsort(pv)
    ranked = pv[order]
    adjusted = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    restored = np.empty_like(adjusted)
    restored[order] = np.clip(adjusted, 0, 1)
    out[np.flatnonzero(valid)] = restored
    return pd.Series(out, index=values.index)


def sign_p_greater(values: pd.Series | np.ndarray) -> float:
    vals = pd.to_numeric(pd.Series(values), errors="coerce").dropna().to_numpy(dtype=float)
    vals = vals[vals != 0]
    if vals.size == 0:
        return np.nan
    return float(stats.binomtest(int(np.sum(vals > 0)), int(vals.size), 0.5, alternative="greater").pvalue)


def bootstrap_median_ci(values: pd.Series | np.ndarray, seed_offset: int = 0) -> tuple[float, float]:
    vals = pd.to_numeric(pd.Series(values), errors="coerce").dropna().to_numpy(dtype=float)
    if vals.size < 2:
        return np.nan, np.nan
    rng = np.random.default_rng(205920 + seed_offset)
    draws = rng.choice(vals, size=(N_BOOTSTRAP, vals.size), replace=True)
    medians = np.median(draws, axis=1)
    return float(np.quantile(medians, 0.025)), float(np.quantile(medians, 0.975))


def cliffs_delta(a: np.ndarray, b: np.ndarray) -> float:
    if len(a) == 0 or len(b) == 0:
        return np.nan
    comparisons = np.subtract.outer(a, b)
    return float((np.sum(comparisons > 0) - np.sum(comparisons < 0)) / comparisons.size)


def load_inputs() -> tuple[pd.DataFrame, ...]:
    deltas = pd.read_csv(DATASET_DELTAS, sep="\t", low_memory=False)
    deltas = deltas[deltas["dataset"].isin(STRICT_DATASETS)].copy()
    deltas["species"] = deltas["dataset"].map(lambda x: STRICT_DATASETS[x][0])
    deltas["region"] = deltas["dataset"].map(lambda x: STRICT_DATASETS[x][1])
    genes = pd.read_csv(GENE_SUMMARY, sep="\t", low_memory=False)
    tiers = pd.read_csv(CANDIDATE_TIERS, sep="\t")
    configuration = pd.read_csv(CONFIGURATION, sep="\t")
    module_summary = pd.read_csv(MODULE_SUMMARY, sep="\t")
    module_gene_scores = pd.read_csv(MODULE_GENE_SCORES, sep="\t")
    module_gene_sets = pd.read_csv(MODULE_GENE_SETS, sep="\t")
    return deltas, genes, tiers, configuration, module_summary, module_gene_scores, module_gene_sets


def build_dataset_level_configuration(configuration: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    dataset_rows = (
        configuration.groupby(["dataset", "core_branch"], as_index=False)
        .agg(
            n_nested_contrasts=("delta_configuration_score", "size"),
            median_delta_configuration_score=("delta_configuration_score", "median"),
            median_delta_construction_balance=("delta_downstream_construction_balance", "median"),
            median_delta_regional_fate_balance=("delta_regional_fate_balance", "median"),
            positive_nested_fraction=("configuration_positive", "mean"),
        )
    )
    dataset_rows["species"] = dataset_rows["dataset"].map(
        lambda x: STRICT_DATASETS.get(x, ("unmapped", "unmapped"))[0]
    )
    dataset_rows["region"] = dataset_rows["dataset"].map(
        lambda x: STRICT_DATASETS.get(x, ("unmapped", "unmapped"))[1]
    )
    dataset_rows["dataset_median_positive"] = dataset_rows["median_delta_configuration_score"].gt(0)

    summaries: list[dict[str, object]] = []
    groups: list[tuple[str, pd.DataFrame]] = [("all", dataset_rows)]
    groups.extend((f"region:{name}", sub) for name, sub in dataset_rows.groupby("region"))
    groups.extend((f"species:{name}", sub) for name, sub in dataset_rows.groupby("species"))
    for offset, (group_id, sub) in enumerate(groups):
        vals = sub["median_delta_configuration_score"]
        lo, hi = bootstrap_median_ci(vals, offset)
        summaries.append(
            {
                "summary_group": group_id,
                "n_independent_datasets": int(len(sub)),
                "n_positive_dataset_medians": int(vals.gt(0).sum()),
                "median_dataset_delta": float(vals.median()),
                "bootstrap_95ci_low": lo,
                "bootstrap_95ci_high": hi,
                "exact_sign_p_greater": sign_p_greater(vals),
                "datasets": ",".join(sorted(sub["dataset"].astype(str))),
            }
        )
    summary = pd.DataFrame(summaries)
    dataset_rows.to_csv(OUT_CONFIG_DATASETS, sep="\t", index=False)
    summary.to_csv(OUT_CONFIG_SUMMARY, sep="\t", index=False)
    return dataset_rows, summary


def build_module_level_inference(module_summary: pd.DataFrame) -> pd.DataFrame:
    out = module_summary[
        [
            "module_id",
            "module_label",
            "module_family",
            "n_genes_defined",
            "median_overall_convergence_delta",
        ]
    ].copy()
    out["inference_group"] = np.where(
        out["module_family"].eq("downstream_circuit_morphology"), "downstream", "upstream_or_niche"
    )
    downstream = out.loc[out["inference_group"].eq("downstream"), "median_overall_convergence_delta"].to_numpy()
    upstream = out.loc[
        out["inference_group"].eq("upstream_or_niche"), "median_overall_convergence_delta"
    ].to_numpy()
    p_value = float(stats.mannwhitneyu(downstream, upstream, alternative="greater", method="exact").pvalue)
    comparison = pd.DataFrame(
        [
            {
                "module_id": "downstream_vs_upstream_module_level",
                "module_label": "Downstream versus upstream/niche modules",
                "module_family": "module_level_comparison",
                "n_genes_defined": int(out["n_genes_defined"].sum()),
                "median_overall_convergence_delta": float(np.median(downstream) - np.median(upstream)),
                "inference_group": "comparison",
                "n_downstream_modules": int(len(downstream)),
                "n_upstream_modules": int(len(upstream)),
                "downstream_median": float(np.median(downstream)),
                "upstream_median": float(np.median(upstream)),
                "exact_mannwhitney_p_greater": p_value,
                "cliffs_delta": cliffs_delta(downstream, upstream),
                "interpretation": "descriptive_directional_not_conventionally_significant_at_module_level",
            }
        ]
    )
    out["n_downstream_modules"] = np.nan
    out["n_upstream_modules"] = np.nan
    out["downstream_median"] = np.nan
    out["upstream_median"] = np.nan
    out["exact_mannwhitney_p_greater"] = np.nan
    out["cliffs_delta"] = np.nan
    out["interpretation"] = "module_summary"
    result = pd.concat([out, comparison], ignore_index=True, sort=False)
    result.to_csv(OUT_MODULE_INFERENCE, sep="\t", index=False)
    return result


def build_species_stratification(deltas: pd.DataFrame, tiers: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    selected = tiers[tiers["manuscript_tier"].isin([TIER1_LABEL, TIER2_LABEL])][
        ["gene", "manuscript_tier", "mechanism_class"]
    ].drop_duplicates()
    work = deltas.merge(selected, left_on="canonical_gene", right_on="gene", how="inner", suffixes=("", "_tier"))
    records: list[dict[str, object]] = []
    group_cols = ["canonical_gene", "manuscript_tier", "mechanism_class", "species", "screen", "branch_tested"]
    for offset, (keys, sub) in enumerate(work.groupby(group_cols, sort=False)):
        vals = sub["dataset_rank_delta"]
        lo, hi = bootstrap_median_ci(vals, 100 + offset)
        record = dict(zip(group_cols, keys, strict=False))
        record.update(
            {
                "n_datasets": int(sub["dataset"].nunique()),
                "median_dataset_rank_delta": float(vals.median()),
                "positive_dataset_fraction": float(vals.gt(0).mean()),
                "bootstrap_95ci_low": lo,
                "bootstrap_95ci_high": hi,
                "exact_sign_p_greater": sign_p_greater(vals),
                "datasets": ",".join(sorted(sub["dataset"].unique())),
                "descriptive_positive": bool(vals.median() > 0),
            }
        )
        records.append(record)
    stratified = pd.DataFrame(records)

    shared_records: list[dict[str, object]] = []
    pair_cols = ["canonical_gene", "manuscript_tier", "mechanism_class", "species", "screen"]
    for keys, sub in stratified.groupby(pair_cols, sort=False):
        by_branch = sub.set_index("branch_tested")
        if not {"dentate", "cerebellar"}.issubset(by_branch.index):
            continue
        dentate = by_branch.loc["dentate"]
        cerebellar = by_branch.loc["cerebellar"]
        record = dict(zip(pair_cols, keys, strict=False))
        record.update(
            {
                "dentate_n_datasets": int(dentate["n_datasets"]),
                "cerebellar_n_datasets": int(cerebellar["n_datasets"]),
                "dentate_median_delta": float(dentate["median_dataset_rank_delta"]),
                "cerebellar_median_delta": float(cerebellar["median_dataset_rank_delta"]),
                "shared_minimum_median_delta": float(
                    min(dentate["median_dataset_rank_delta"], cerebellar["median_dataset_rank_delta"])
                ),
                "both_branches_descriptively_positive": bool(
                    dentate["median_dataset_rank_delta"] > 0
                    and cerebellar["median_dataset_rank_delta"] > 0
                ),
                "inference_note": (
                    "replicated_within_species"
                    if min(dentate["n_datasets"], cerebellar["n_datasets"]) >= 2
                    else "descriptive_single_dataset_cerebellar_arm"
                ),
            }
        )
        shared_records.append(record)
    shared = pd.DataFrame(shared_records)
    stratified.to_csv(OUT_SPECIES, sep="\t", index=False)
    shared.to_csv(OUT_SPECIES_SHARED, sep="\t", index=False)
    return stratified, shared


def build_lodo(deltas: pd.DataFrame, tiers: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    selected = tiers[tiers["manuscript_tier"].isin([TIER1_LABEL, TIER2_LABEL])][
        ["gene", "manuscript_tier"]
    ].drop_duplicates()
    work = deltas.merge(selected, left_on="canonical_gene", right_on="gene", how="inner")
    records: list[dict[str, object]] = []
    for (gene, tier, screen, branch), sub in work.groupby(
        ["canonical_gene", "manuscript_tier", "screen", "branch_tested"], sort=False
    ):
        datasets = sorted(sub["dataset"].unique())
        baseline_vals = sub["dataset_rank_delta"]
        records.append(
            {
                "canonical_gene": gene,
                "manuscript_tier": tier,
                "screen": screen,
                "branch_tested": branch,
                "held_out_dataset": "none",
                "n_datasets_remaining": int(len(datasets)),
                "median_dataset_rank_delta": float(baseline_vals.median()),
                "positive_dataset_fraction": float(baseline_vals.gt(0).mean()),
                "robust_directional_support": bool(
                    len(datasets) >= 2
                    and baseline_vals.median() > 0
                    and baseline_vals.gt(0).mean() >= 0.75
                ),
            }
        )
        for held_out in datasets:
            remaining = sub[sub["dataset"].ne(held_out)]
            vals = remaining["dataset_rank_delta"]
            if remaining["dataset"].nunique() < 2:
                continue
            records.append(
                {
                    "canonical_gene": gene,
                    "manuscript_tier": tier,
                    "screen": screen,
                    "branch_tested": branch,
                    "held_out_dataset": held_out,
                    "n_datasets_remaining": int(remaining["dataset"].nunique()),
                    "median_dataset_rank_delta": float(vals.median()),
                    "positive_dataset_fraction": float(vals.gt(0).mean()),
                    "robust_directional_support": bool(
                        vals.median() > 0 and vals.gt(0).mean() >= 0.75
                    ),
                }
            )
    lodo = pd.DataFrame(records)
    held = lodo[lodo["held_out_dataset"].ne("none")]
    summary = (
        held.groupby(["canonical_gene", "manuscript_tier"], as_index=False)
        .agg(
            n_leave_one_dataset_tests=("robust_directional_support", "size"),
            n_stable_tests=("robust_directional_support", "sum"),
            lodo_stability_fraction=("robust_directional_support", "mean"),
            minimum_lodo_median_delta=("median_dataset_rank_delta", "min"),
            minimum_lodo_positive_fraction=("positive_dataset_fraction", "min"),
        )
    )
    summary["stable_all_testable_leave_outs"] = summary["lodo_stability_fraction"].eq(1.0)
    lodo.to_csv(OUT_LODO, sep="\t", index=False)
    summary.to_csv(OUT_LODO_SUMMARY, sep="\t", index=False)
    return lodo, summary


def mechanism_group(value: object) -> str:
    text = str(value or "").lower()
    if re.search(r"synap|channel|excit|calcium|glutamate|gaba|vesicle|receptor", text):
        return "synaptic_excitability"
    if re.search(r"morph|cytoskel|guidance|adhesion|neurite", text):
        return "morphology_guidance"
    if re.search(r"regulat|transcription|rna_processing|chromatin", text):
        return "regulatory"
    return "broad_other"


def build_matched_nulls(genes: pd.DataFrame, tiers: pd.DataFrame) -> pd.DataFrame:
    pool = genes.copy()
    pool["function_group"] = pool["selected_mechanism_class"].where(
        pool["selected_mechanism_class"].notna(), pool["genome_mechanism_class"]
    ).map(mechanism_group)
    tier_map = tiers[["gene", "manuscript_tier", "mechanism_class"]].drop_duplicates()
    tier_map["function_group"] = tier_map["mechanism_class"].map(mechanism_group)
    excluded = set(tier_map["gene"])
    pool = pool[
        pool["canonical_gene"].notna()
        & pool["ortholog_symbol_class"].eq("same_symbol")
        & ~pool["canonical_gene"].isin(excluded)
    ].copy()

    def build_match_options(targets: pd.DataFrame) -> list[np.ndarray]:
        options: list[np.ndarray] = []
        for _, target in targets.iterrows():
            candidates = pool[
                pool["formal_n_available_branches"].eq(target["formal_n_available_branches"])
                & pool["function_group"].eq(target["function_group"])
            ].copy()
            if len(candidates) < 10:
                candidates = pool[
                    pool["formal_n_available_branches"].eq(target["formal_n_available_branches"])
                ].copy()
            target_detection = float(target["formal_min_candidate_detection"])
            candidates["match_distance"] = (
                pd.to_numeric(candidates["formal_min_candidate_detection"], errors="coerce")
                - target_detection
            ).abs()
            candidates = candidates.nsmallest(min(100, len(candidates)), "match_distance")
            options.append(candidates["canonical_gene"].astype(str).to_numpy())
        return options

    def draw_set(options: list[np.ndarray], rng: np.random.Generator) -> list[str]:
        chosen: list[str] = []
        for candidates in options:
            available = np.array([gene for gene in candidates if gene not in chosen], dtype=object)
            if available.size == 0:
                available = candidates
            choice = str(rng.choice(available))
            chosen.append(choice)
        return chosen

    records: list[dict[str, object]] = []
    for label, tier_labels in [
        ("Tier1", [TIER1_LABEL]),
        ("Tier1_plus_Tier2", [TIER1_LABEL, TIER2_LABEL]),
    ]:
        targets = tier_map[tier_map["manuscript_tier"].isin(tier_labels)].merge(
            genes[
                [
                    "canonical_gene",
                    "formal_n_available_branches",
                    "formal_min_candidate_detection",
                    "formal_rank_priority_score",
                    "formal_n_nominal_branches",
                    "formal_nominal_shared_both_screens",
                    "median_dataset_rank_delta_full_matrix_cerebellar",
                    "median_dataset_rank_delta_full_matrix_dentate",
                    "median_dataset_rank_delta_selected_cerebellar",
                    "median_dataset_rank_delta_selected_dentate",
                ]
            ],
            left_on="gene",
            right_on="canonical_gene",
            how="left",
        )
        targets["function_group"] = targets["mechanism_class"].map(mechanism_group)

        def metrics(frame: pd.DataFrame) -> tuple[float, float, float]:
            branch_cols = [
                "median_dataset_rank_delta_full_matrix_cerebellar",
                "median_dataset_rank_delta_full_matrix_dentate",
                "median_dataset_rank_delta_selected_cerebellar",
                "median_dataset_rank_delta_selected_dentate",
            ]
            min_delta = frame[branch_cols].min(axis=1, skipna=True)
            return (
                float(frame["formal_rank_priority_score"].mean()),
                float(frame["formal_nominal_shared_both_screens"].fillna(False).mean()),
                float(min_delta.mean()),
            )

        observed = metrics(targets)
        match_options = build_match_options(targets)
        gene_lookup = genes.drop_duplicates("canonical_gene").set_index("canonical_gene")
        null_metrics = np.empty((N_NULL, 3), dtype=float)
        for i in range(N_NULL):
            selected_genes = draw_set(match_options, RNG)
            null_frame = gene_lookup.loc[selected_genes].reset_index()
            null_metrics[i, :] = metrics(null_frame)
        metric_names = [
            "mean_formal_rank_priority_score",
            "fraction_nominal_shared_both_screens",
            "mean_minimum_branch_median_delta",
        ]
        for j, metric_name in enumerate(metric_names):
            null_values = null_metrics[:, j]
            records.append(
                {
                    "candidate_set": label,
                    "n_genes": int(len(targets)),
                    "metric": metric_name,
                    "observed_value": observed[j],
                    "null_median": float(np.nanmedian(null_values)),
                    "null_q025": float(np.nanquantile(null_values, 0.025)),
                    "null_q975": float(np.nanquantile(null_values, 0.975)),
                    "empirical_p_greater": float(
                        (1 + np.sum(null_values >= observed[j])) / (N_NULL + 1)
                    ),
                    "interpretation": "selection_bias_sensitivity_not_independent_validation",
                    "matching": "same_available_branch_count;function_group_when_possible;nearest_detection",
                }
            )
    out = pd.DataFrame(records)
    out.to_csv(OUT_NULL, sep="\t", index=False)
    return out


def build_module_sensitivity(
    module_gene_scores: pd.DataFrame, module_gene_sets: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    records: list[dict[str, object]] = []
    for (module_id, module_label, module_family), sub in module_gene_scores.groupby(
        ["module_id", "module_label", "module_family"], sort=False
    ):
        vals = pd.to_numeric(sub["overall_convergence_delta"], errors="coerce")
        baseline = float(vals.median())
        for gene in sub["gene"]:
            leave = pd.to_numeric(
                sub.loc[sub["gene"].ne(gene), "overall_convergence_delta"], errors="coerce"
            )
            records.append(
                {
                    "module_id": module_id,
                    "module_label": module_label,
                    "module_family": module_family,
                    "held_out_gene": gene,
                    "n_genes_remaining": int(leave.notna().sum()),
                    "baseline_median_convergence_delta": baseline,
                    "leave_one_gene_out_median_delta": float(leave.median()),
                    "sign_preserved": bool(np.sign(leave.median()) == np.sign(baseline)),
                }
            )
    sensitivity = pd.DataFrame(records)

    modules = {
        module: set(sub["canonical_gene"].dropna().astype(str))
        for module, sub in module_gene_sets.groupby("module_id")
    }
    overlap_records: list[dict[str, object]] = []
    for module_a in modules:
        for module_b in modules:
            if module_a >= module_b:
                continue
            intersection = modules[module_a] & modules[module_b]
            union = modules[module_a] | modules[module_b]
            overlap_records.append(
                {
                    "module_a": module_a,
                    "module_b": module_b,
                    "n_a": len(modules[module_a]),
                    "n_b": len(modules[module_b]),
                    "n_overlap": len(intersection),
                    "jaccard": len(intersection) / len(union) if union else np.nan,
                    "overlap_genes": ",".join(sorted(intersection)),
                }
            )
    overlap = pd.DataFrame(overlap_records)
    sensitivity.to_csv(OUT_MODULE_SENSITIVITY, sep="\t", index=False)
    overlap.to_csv(OUT_MODULE_OVERLAP, sep="\t", index=False)
    return sensitivity, overlap


def split_markers(value: object) -> set[str]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return set()
    return {token.strip().upper() for token in str(value).split(",") if token.strip()}


def ccf_frequency(value: object, acronym: str) -> float:
    if value is None:
        return 0.0
    pattern = rf"(?:^|,){re.escape(acronym)}:([0-9.]+)"
    match = re.search(pattern, str(value))
    return float(match.group(1)) if match else 0.0


def allen_population(row: pd.Series) -> str | None:
    subclass = str(row.get("subclass_label") or "")
    class_label = str(row.get("class_label") or "")
    nt = str(row.get("nt_type_label") or "")
    ccf = row.get("CCF_acronym.freq")
    if subclass == "DG Glut":
        return "Dentate granule"
    if subclass == "CB Granule Glut":
        return "Cerebellar granule"
    if class_label == "OB-IMN GABA" and nt == "GABA" and ccf_frequency(ccf, "MOBgr") >= 0.20:
        return "Olfactory-bulb granule-layer GABA proxy"
    if subclass.startswith("L4") and nt == "Glut":
        return "Cortical L4 excitatory proxy"
    if subclass in {"CA1-ProS Glut", "CA3 Glut"}:
        return "Hippocampal pyramidal comparator"
    if subclass == "CBX Purkinje Gaba":
        return "Purkinje comparator"
    return None


def load_allen_clusters() -> pd.DataFrame:
    workbook = load_workbook(ALLEN_XLSX, read_only=True, data_only=True)
    sheet = workbook["cluster_annotation"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    frame = pd.DataFrame(rows, columns=header)
    frame["comparator_population"] = frame.apply(allen_population, axis=1)
    frame = frame[frame["comparator_population"].notna()].copy()
    marker_cols = [
        "cluster.markers.combo",
        "merfish.markers.combo",
        "cluster.TF.markers.combo",
        "cluster.markers.combo (within subclass)",
    ]
    frame["marker_set"] = frame.apply(
        lambda row: set().union(*(split_markers(row[col]) for col in marker_cols)), axis=1
    )
    frame["n_unique_markers"] = frame["marker_set"].map(len)
    export = frame[
        [
            "cluster_id",
            "cluster_id_label",
            "supertype_label",
            "subclass_label",
            "class_label",
            "neighborhood",
            "anatomical_annotation",
            "CCF_acronym.freq",
            "nt_type_label",
            "comparator_population",
            "n_unique_markers",
        ]
    ].copy()
    export["combined_markers"] = frame["marker_set"].map(lambda x: ",".join(sorted(x)))
    export.to_csv(OUT_ALLEN_CLUSTERS, sep="\t", index=False)
    return frame


def build_allen_comparator(
    clusters: pd.DataFrame, module_gene_sets: pd.DataFrame, tiers: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    module_lookup = {
        module: set(sub["canonical_gene"].dropna().astype(str).str.upper())
        for module, sub in module_gene_sets.groupby("module_id")
    }
    module_labels = module_gene_sets.drop_duplicates("module_id").set_index("module_id")[
        ["module_label", "module_family"]
    ]
    module_records: list[dict[str, object]] = []
    for population, sub in clusters.groupby("comparator_population", sort=False):
        population_markers = set().union(*sub["marker_set"].tolist())
        for module_id, genes in module_lookup.items():
            per_cluster_hits = sub["marker_set"].map(lambda markers: len(markers & genes))
            per_cluster_fraction = [
                len(markers & genes) / len(markers) if markers else 0.0 for markers in sub["marker_set"]
            ]
            overlap = population_markers & genes
            module_records.append(
                {
                    "comparator_population": population,
                    "n_clusters": int(len(sub)),
                    "module_id": module_id,
                    "module_label": module_labels.loc[module_id, "module_label"],
                    "module_family": module_labels.loc[module_id, "module_family"],
                    "module_size": int(len(genes)),
                    "n_module_genes_seen_as_markers": int(len(overlap)),
                    "module_gene_coverage": float(len(overlap) / len(genes)) if genes else np.nan,
                    "fraction_clusters_with_module_marker": float(per_cluster_hits.gt(0).mean()),
                    "mean_fraction_of_cluster_markers_in_module": float(np.mean(per_cluster_fraction)),
                    "marker_overlap_genes": ",".join(sorted(overlap)),
                    "evidence_scope": "adult_Allen_taxonomy_marker_annotation_not_expression_matrix",
                }
            )
    module_out = pd.DataFrame(module_records)

    candidate_records: list[dict[str, object]] = []
    candidate_rows = tiers[tiers["manuscript_tier"].isin([TIER1_LABEL, TIER2_LABEL])]
    for population, sub in clusters.groupby("comparator_population", sort=False):
        for _, candidate in candidate_rows.iterrows():
            gene = str(candidate["gene"]).upper()
            hits = sub["marker_set"].map(lambda markers: gene in markers)
            candidate_records.append(
                {
                    "comparator_population": population,
                    "n_clusters": int(len(sub)),
                    "gene": gene,
                    "manuscript_tier": candidate["manuscript_tier"],
                    "mechanism_class": candidate["mechanism_class"],
                    "n_clusters_with_gene_as_marker": int(hits.sum()),
                    "cluster_marker_prevalence": float(hits.mean()),
                    "is_population_exclusive_among_tested_groups": False,
                }
            )
    candidate_out = pd.DataFrame(candidate_records)
    for gene, idx in candidate_out.groupby("gene").groups.items():
        positive_groups = int(candidate_out.loc[idx, "n_clusters_with_gene_as_marker"].gt(0).sum())
        candidate_out.loc[idx, "n_detected_population_groups"] = positive_groups
        candidate_out.loc[idx, "detected_in_any_tested_group"] = positive_groups > 0
        candidate_out.loc[idx, "is_population_exclusive_among_tested_groups"] = (
            positive_groups == 1 if positive_groups > 0 else np.nan
        )

    module_out.to_csv(OUT_ALLEN_MODULES, sep="\t", index=False)
    candidate_out.to_csv(OUT_ALLEN_CANDIDATES, sep="\t", index=False)
    return module_out, candidate_out


def build_figure(
    config_datasets: pd.DataFrame,
    species_shared: pd.DataFrame,
    lodo_summary: pd.DataFrame,
    allen_modules: pd.DataFrame,
) -> None:
    sns.set_theme(style="whitegrid", font_scale=1.0)
    fig, axes = plt.subplots(2, 2, figsize=(15, 11), constrained_layout=True)

    ax = axes[0, 0]
    order = config_datasets.sort_values("median_delta_configuration_score")["dataset"]
    plot = config_datasets.set_index("dataset").loc[order]
    colors = plot["region"].map({"dentate": "#2B6CB0", "cerebellum": "#D97706"})
    ax.barh(plot.index, plot["median_delta_configuration_score"], color=colors)
    ax.axvline(0, color="black", linewidth=1)
    ax.set_xlabel("Dataset-median configuration delta")
    ax.set_ylabel("")
    ax.set_title("a  Dataset-level configuration support")

    ax = axes[0, 1]
    mouse = species_shared[
        species_shared["species"].eq("mouse")
        & species_shared["screen"].eq("full_matrix")
    ].copy()
    if mouse.empty:
        mouse = species_shared[species_shared["species"].eq("mouse")].copy()
    mouse = mouse.sort_values(["manuscript_tier", "shared_minimum_median_delta"], ascending=[True, False])
    palette = mouse["manuscript_tier"].map({TIER1_LABEL: "#1B9E77", TIER2_LABEL: "#7570B3"})
    ax.barh(mouse["canonical_gene"], mouse["shared_minimum_median_delta"], color=palette)
    ax.axvline(0, color="black", linewidth=1)
    ax.set_xlabel("Weaker-branch median rank delta")
    ax.set_ylabel("")
    ax.set_title("b  Mouse-only shared candidate support")
    ax.invert_yaxis()

    ax = axes[1, 0]
    matrix = allen_modules.pivot(
        index="module_label",
        columns="comparator_population",
        values="fraction_clusters_with_module_marker",
    )
    preferred_columns = [
        "Dentate granule",
        "Cerebellar granule",
        "Olfactory-bulb granule-layer GABA proxy",
        "Cortical L4 excitatory proxy",
        "Hippocampal pyramidal comparator",
        "Purkinje comparator",
    ]
    matrix = matrix[[col for col in preferred_columns if col in matrix.columns]]
    sns.heatmap(matrix, cmap="viridis", vmin=0, vmax=1, annot=True, fmt=".2f", ax=ax, cbar_kws={"label": "Cluster fraction"})
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.set_title("c  Allen adult taxonomy marker support")
    ax.tick_params(axis="x", rotation=35)

    ax = axes[1, 1]
    summary = lodo_summary.sort_values(["manuscript_tier", "lodo_stability_fraction"])
    palette = summary["manuscript_tier"].map({TIER1_LABEL: "#1B9E77", TIER2_LABEL: "#7570B3"})
    ax.barh(summary["canonical_gene"], summary["lodo_stability_fraction"], color=palette)
    ax.set_xlim(0, 1.03)
    ax.set_xlabel("Fraction of testable leave-one-dataset analyses stable")
    ax.set_ylabel("")
    ax.set_title("d  Candidate leave-one-dataset stability")
    ax.invert_yaxis()

    fig.savefig(OUT_FIGURE, dpi=300, bbox_inches="tight")
    plt.close(fig)


def build_report(
    config_summary: pd.DataFrame,
    module_inference: pd.DataFrame,
    species_shared: pd.DataFrame,
    lodo_summary: pd.DataFrame,
    null_summary: pd.DataFrame,
    module_sensitivity: pd.DataFrame,
    allen_clusters: pd.DataFrame,
    allen_candidates: pd.DataFrame,
) -> None:
    overall = config_summary[config_summary["summary_group"].eq("all")].iloc[0]
    module_test = module_inference[
        module_inference["module_id"].eq("downstream_vs_upstream_module_level")
    ].iloc[0]
    mouse = species_shared[species_shared["species"].eq("mouse")]
    mouse_positive = int(mouse["both_branches_descriptively_positive"].sum())
    mouse_total = int(len(mouse))
    stable_tier1 = lodo_summary[lodo_summary["manuscript_tier"].eq(TIER1_LABEL)]
    tier1_stable = int(stable_tier1["stable_all_testable_leave_outs"].sum())
    sensitivity_summary = (
        module_sensitivity.groupby("module_label")["sign_preserved"].mean().sort_values()
    )
    group_counts = allen_clusters["comparator_population"].value_counts()
    candidate_detection = allen_candidates.groupby("gene").agg(
        detected=("detected_in_any_tested_group", "max"),
        n_groups=("n_detected_population_groups", "max"),
    )
    n_candidate_detected = int(candidate_detection["detected"].sum())
    lines = [
        "# Dataset-level robustness and expanded comparator analysis",
        "",
        "## Main conclusions",
        "",
        (
            f"- Collapsing the 63 nested configuration contrasts to independent dataset medians retained "
            f"positive support in {int(overall['n_positive_dataset_medians'])}/{int(overall['n_independent_datasets'])} "
            f"datasets. The dataset-level median delta was {overall['median_dataset_delta']:.3f} "
            f"(bootstrap 95% CI {overall['bootstrap_95ci_low']:.3f} to {overall['bootstrap_95ci_high']:.3f}; "
            f"one-sided exact sign p={overall['exact_sign_p_greater']:.4g})."
        ),
        (
            f"- At the five-module level, downstream modules retained a larger median convergence delta "
            f"than upstream/niche modules (difference={module_test['median_overall_convergence_delta']:.3f}; "
            f"Cliff's delta={module_test['cliffs_delta']:.3f}), but the exact one-sided Mann-Whitney "
            f"p={module_test['exact_mannwhitney_p_greater']:.3g}. The result is therefore directional and "
            "descriptive at the module level, not a conventionally significant five-module test."
        ),
        (
            f"- In mouse-only analyses, {mouse_positive}/{mouse_total} candidate-by-screen comparisons "
            "had positive median rank deltas in both dentate and cerebellar branches. Human support remains "
            "descriptive where the cerebellar arm contains one dataset."
        ),
        (
            f"- {tier1_stable}/{len(stable_tier1)} Tier 1 genes retained the directional support rule in "
            "every testable leave-one-dataset analysis."
        ),
        (
            "- Matched-null results are reported as a selection-bias sensitivity analysis. They cannot be "
            "treated as independent validation because the candidate tiers were selected from the same rank-meta evidence."
        ),
        (
            f"- Allen taxonomy marker annotations contributed {len(allen_clusters)} clusters across "
            f"{len(group_counts)} populations. The analysis explicitly labels olfactory-bulb granule-layer "
            "GABA and cortical L4 excitatory cells as proxies because taxonomy markers do not establish morphology."
        ),
        (
            f"- Only {n_candidate_detected}/{allen_candidates['gene'].nunique()} tested Tier 1/2 genes appeared "
            "in the sparse taxonomy-marker lists for these selected populations. The marker annotations therefore "
            "cannot validate candidate expression or specificity; they mainly define comparator identities and show "
            "that the candidate set is not a set of discriminative adult taxonomy markers."
        ),
        "",
        "## Allen comparator cluster counts",
        "",
    ]
    lines.extend(f"- {name}: {count}" for name, count in group_counts.items())
    lines.extend(
        [
            "",
            "## Module leave-one-gene-out sign stability",
            "",
        ]
    )
    lines.extend(f"- {name}: {value:.1%}" for name, value in sensitivity_summary.items())
    lines.extend(
        [
            "",
            "## Interpretation boundary",
            "",
            "These analyses support a narrower conclusion than the rejected manuscript. Dentate and cerebellar granule-cell lineages preserve different regional identities while partially reusing broad postmitotic assembly and maturation machinery. The shared signal is not exclusive to granule cells, the module-level contrast is underpowered, and the Allen extension is an adult marker-annotation sensitivity layer rather than a matched developmental expression analysis.",
            "",
            "## Outputs",
            "",
            f"- `{OUT_CONFIG_SUMMARY.relative_to(ROOT)}`",
            f"- `{OUT_MODULE_INFERENCE.relative_to(ROOT)}`",
            f"- `{OUT_SPECIES_SHARED.relative_to(ROOT)}`",
            f"- `{OUT_LODO_SUMMARY.relative_to(ROOT)}`",
            f"- `{OUT_NULL.relative_to(ROOT)}`",
            f"- `{OUT_MODULE_SENSITIVITY.relative_to(ROOT)}`",
            f"- `{OUT_ALLEN_MODULES.relative_to(ROOT)}`",
            f"- `{OUT_ALLEN_CANDIDATES.relative_to(ROOT)}`",
            f"- `{OUT_FIGURE.relative_to(ROOT)}`",
        ]
    )
    OUT_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    (
        deltas,
        genes,
        tiers,
        configuration,
        module_summary,
        module_gene_scores,
        module_gene_sets,
    ) = load_inputs()
    config_datasets, config_summary = build_dataset_level_configuration(configuration)
    module_inference = build_module_level_inference(module_summary)
    _, species_shared = build_species_stratification(deltas, tiers)
    _, lodo_summary = build_lodo(deltas, tiers)
    null_summary = build_matched_nulls(genes, tiers)
    module_sensitivity, _ = build_module_sensitivity(module_gene_scores, module_gene_sets)
    allen_clusters = load_allen_clusters()
    allen_modules, allen_candidates = build_allen_comparator(allen_clusters, module_gene_sets, tiers)
    build_figure(config_datasets, species_shared, lodo_summary, allen_modules)
    build_report(
        config_summary,
        module_inference,
        species_shared,
        lodo_summary,
        null_summary,
        module_sensitivity,
        allen_clusters,
        allen_candidates,
    )
    print(f"Wrote {OUT_REPORT.relative_to(ROOT)}")
    print(f"Wrote {OUT_FIGURE.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
